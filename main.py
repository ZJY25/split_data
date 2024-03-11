import os
import openpyxl
from openpyxl.utils import get_column_letter
import re

path = r"D:\projects\split_data\split_data"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')
# print(workbook.sheetnames)  # 查看work book的表格

sheet = workbook['test']
# print(sheet)

max_row = sheet.max_row
print("max_row", max_row)  # 获取最大行数

column_item = sheet['A1':'A%d' % max_row]
print("column_item", column_item)

business_type = ["88", "118", "158", "188", "228", "288", "388", "39", "79", "119", "229",
                 "单移", "双百", "副卡", "全能卡99", "全能卡129", "全能卡169", "全能卡229",
                 "高互宽回网", "有效回网", "高清回网", "标清回网", "互动回网", "宽带回网",
                 "有效预离网回网", "宽带预离网回网", "高互宽预离网回网"]


for key, value in enumerate(business_type):
    print(str(key) + ":" + value)

for i in range(5, 10):
    column = 2

    all_string = sheet.cell(row=i, column=1).value
    print(str(i), ":", all_string)

    # ((\d*) | ([(（][\u4e00-\u9fa5]+[)）])*)
    names = re.findall('发展人[:：,，][（）\d\u4e00-\u9fa5]+', all_string)
    for name in names:
        name = re.findall("(?<=[:：,，]).*$", name)
        name_str = name[0]
        column_letter = get_column_letter(column)
        print(column_letter + str(i) + ":" + name_str)
        sheet[column_letter + str(i)] = name_str
        column += 1

    column = 4

    date = re.findall('\d+月\d+日', all_string)
    date_str = date[0]
    column_letter = get_column_letter(column)
    print(column_letter + str(i) + ":" + date_str)
    sheet[column_letter + str(i)] = date_str
    column += 1

    business = re.findall('(?<=业务[:：])[\d\u4e00-\u9fa5]+(?=[,，])', all_string)
    business_str = business[0]
    column_letter = get_column_letter(column)
    print(column_letter + str(i) + ":" + business_str)
    sheet[column_letter + str(i)] = business_str
    column += 1

# workbook.save('test.xlsx')