import os
import openpyxl
from openpyxl.utils import get_column_letter


class ExcelProcess(object):

    def __init__(self, path, filename):
        os.chdir(path)  # 修改工作路径
        self.workbook = openpyxl.load_workbook(filename)
        self.max_row = None
        self.sheet = None

    def load_excel_sheet(self, my_table_name):
        self.sheet = self.workbook[my_table_name]
        return self.sheet

    def get_max_row(self):
        self.max_row = self.sheet.max_row
        return self.max_row

    def save_excel(self, filename):
        self.workbook.save(filename)

    def write_cell(self, i, j, string):
        if isinstance(j, int):
            cell_id = get_column_letter(j) + str(i)
        else:
            cell_id = j + str(i)
        self.sheet[cell_id] = string

    def get_cell(self, i, j):
        if isinstance(j, int):
            cell_id = get_column_letter(j) + str(i)
        else:
            cell_id = j + str(i)
        return self.sheet[cell_id].value

    def write_G(self, row, string):
        cell_id = "G" + str(row)
        self.sheet[cell_id] = string

    def write_H(self, row, string):
        cell_id = "H" + str(row)
        self.sheet[cell_id] = string

    def write_valid_add(self, row):
        cell_id = "I" + str(row)
        self.sheet[cell_id] = "有效新增"

    def write_interactive_add(self, row):
        cell_id = "J" + str(row)
        self.sheet[cell_id] = "互动新增"

    def write_broadband_add(self, row):
        cell_id = "K" + str(row)
        self.sheet[cell_id] = "宽带新增"
